# 入居管理表アプリ（自動切替版）
# - 文字PDFなら：AIなし（pdfの構造解析）で高速・無料
# - スキャンPDFなら：AI（画像読み取り→JSON）で対応
# - 入力PDFごとに自動判定してルート切替
# - 既存のExcel出力・Pxx付替え・備考ユニーク化・基準額（最頻値）などは維持

import streamlit as st
import io
import json
import asyncio
import base64
import re
import logging
from datetime import datetime
from pathlib import Path
from collections import Counter
import subprocess

import pdfplumber
from PIL import Image
import fitz  # PyMuPDF

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

# ===== ログ =====
if not logging.getLogger().hasHandlers():
    log_filename = f"app_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    logging.basicConfig(
        level=logging.INFO,
        format='[%(asctime)s] %(levelname)s: %(message)s',
        handlers=[logging.FileHandler(log_filename, encoding='utf-8'),
                  logging.StreamHandler()]
    )
logger = logging.getLogger(__name__)

# ===== サイドバー：ブランチ/コミット表示 =====
try:
    BRANCH = subprocess.check_output(["git","rev-parse","--abbrev-ref","HEAD"], text=True).strip()
    COMMIT = subprocess.check_output(["git","rev-parse","--short","HEAD"], text=True).strip()
    st.sidebar.info(f"branch: {BRANCH}\ncommit: {COMMIT}")
except Exception as e:
    st.sidebar.warning(f"git情報を取得できませんでした: {e}")

# ===== Excel禁止文字の除去ユーティリティ =====
def xls_clean(v):
    if v is None:
        return None
    s = str(v)
    s = ILLEGAL_CHARACTERS_RE.sub("", s)  # openpyxl の禁止文字（制御文字など）を除去
    return s

# ========== 文字列/数値ユーティリティ ==========
def _normalize_cell(s):
    if s is None: return ""
    s = str(s).replace("\x00","").strip()   # NULL 早期除去
    return s

def _to_int_like(s):
    s = _normalize_cell(s).replace(",", "").replace("¥", "").replace("円", "")
    if s == "": return 0
    try: return int(float(s))
    except: return 0

def extract_text_with_pdfplumber(pdf_bytes: bytes) -> str:
    texts = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            texts.append(page.extract_text() or "")
    return "\n".join(texts)

def extract_month_from_filename(filename: str) -> str:
    m = re.search(r"(\d{4})年(\d{1,2})月", filename)
    return f"{m.group(1)}-{m.group(2).zfill(2)}" if m else "unknown"

def normalize_room(s: str) -> str:
    """ 0101 / 0205 / 0303 / P01 などへ正規化 """
    if not s:
        return s
    s = str(s).strip()
    if re.fullmatch(r"P\d{1,2}", s, re.IGNORECASE):
        p = s.upper().replace("P", "")
        return f"P{p.zfill(2)}"
    digits = re.sub(r"\D", "", s)
    if digits:
        return digits.zfill(4)
    return s

def clean_int(v):
    if v is None: return 0
    if isinstance(v, (int, float)): return int(v)
    s = str(v).replace(",", "").strip()
    if s == "": return 0
    try: return int(float(s))
    except: return 0

def month_key(s: str) -> str:
    m = re.match(r"(\d{4})[-/年](\d{1,2})", str(s))
    if not m: return s
    return f"{m.group(1)}-{m.group(2).zfill(2)}"

# 備考をカンマ区切りで保持しつつ、重複を除去して結合
def append_note_unique(current: str, note: str) -> str:
    """
    current: 既存の備考（'a, b' など）
    note   : 追加したい備考（'b' など）
    返り値: 重複を取り除き、順序維持で 'a, b, c' の形にして返す
    """
    if not note:
        return current or ""
    def _tok(s: str):
        return [t for t in re.split(r"[,\u3001]\s*", (s or "")) if t]
    tokens = _tok(current)
    seen = set(tokens)
    n = note.strip()
    if n and n not in seen:
        tokens.append(n)
    return ", ".join(tokens)

# =========================================================
# 自動判定：このPDFは「文字PDF」か？
#   - 最初の数ページの page.chars 合計で簡易に判断
# =========================================================
def is_text_pdf(pdf_bytes: bytes, min_chars: int = 50, sample_pages: int = 3) -> bool:
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            counts = [len(p.chars or []) for p in pdf.pages[:sample_pages]]
            total = sum(counts)
            logger.info(f"[judge] chars_per_page={counts} total={total} min={min_chars}")
            return total >= min_chars
    except Exception as e:
        logger.warning(f"[judge] exception in is_text_pdf: {e}")
        return False

# =========================================================
# 非AIルート（pdf構造解析）—— 既存の pdfplumber-only を関数化
# =========================================================
def extract_income_table_with_pdfplumber(pdf_bytes: bytes, top_margin_px: int = 40, side_margin_px: int = 24):
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            if len(pdf.pages) == 0:
                return None
            page = pdf.pages[1] if len(pdf.pages) >= 2 else pdf.pages[0]
            if not page.chars or len(page.chars) < 10:
                return None
            W, H = page.width, page.height
            px_to_pt = 0.75
            x0 = side_margin_px * px_to_pt
            x1 = W - x0
            y0 = top_margin_px * px_to_pt
            y1 = H - (12 * px_to_pt)
            crop = page.crop((x0, y0, x1, y1))
            lattice = {
                "vertical_strategy":"lines","horizontal_strategy":"lines",
                "snap_tolerance":3,"join_tolerance":3,
                "intersection_x_tolerance":5,"intersection_y_tolerance":5,
                "edge_min_length":30,
            }
            tables = crop.extract_tables(lattice)
            if not tables:
                stream = {
                    "vertical_strategy":"text","horizontal_strategy":"text",
                    "text_x_tolerance":2,"text_y_tolerance":2,
                    "snap_tolerance":3,"join_tolerance":3,
                }
                tables = crop.extract_tables(stream)
            if not tables:
                return None
            table = tables[0]
            cleaned = [[ _normalize_cell(c) for c in row ] for row in table]
            cells = sum(len(r) for r in cleaned)
            nonempty = sum(1 for r in cleaned for c in r if c)
            if cells == 0 or (nonempty / cells) < 0.01:
                return None
            return cleaned
    except Exception as e:
        logger.warning(f"pdfplumber抽出で例外: {e}")
        return None

def parse_income_table_to_records(table_2d, default_month_id: str):
    if not table_2d or len(table_2d) < 2:
        return []
    headers = [re.sub(r"\s+", "", h) for h in table_2d[0]]
    data_rows = table_2d[1:]

    def find_col(*names):
        for n in names:
            if n in headers:
                return headers.index(n)
        return None

    COL = {
        "room":    find_col("部屋","室","号室","室番号"),
        "tenant":  find_col("契約者","賃借人","入居者","名義"),
        "month":   find_col("年／月","年/月","年月"),
        "rent":    find_col("賃料","家賃"),
        "fee":     find_col("共益費","管理費"),
        "parking": find_col("駐車料","Ｐ","Ｐ料金","P料金"),
        "water":   find_col("水道代","水道料"),
        "reikin":  find_col("礼金"),
        "koushin": find_col("更新料"),
        "bikou":   find_col("備考","摘要","特記事項"),
    }

    def at(row, key):
        j = COL.get(key)
        if j is None or j >= len(row): return ""
        return _normalize_cell(row[j])

    def month_from_cell(cell_value: str, fallback: str) -> str:
        s = str(cell_value or "").strip()
        if not s:
            return month_key(fallback)
        m = re.search(r'(\d{2,4})[年/\-\.](\d{1,2})', s)
        if m:
            yy = m.group(1)
            mm = m.group(2).zfill(2)
            if len(yy) == 2:
                yy = "20" + yy
            return f"{yy}-{mm}"
        return month_key(fallback)

    records = []
    for row in data_rows:
        room_raw   = at(row, "room")
        tenant_raw = at(row, "tenant")
        if re.search(r"合\s*計|総\s*計|合計額|総合計", room_raw) or \
           re.search(r"合\s*計|総\s*計|合計額|総合計", tenant_raw):
            continue

        mk = month_from_cell(at(row, "month"), default_month_id)
        rec = {
            "room":        normalize_room(room_raw),
            "tenant":      (tenant_raw or "").strip(),
            "monthly": {
                mk: {
                    "rent":    clean_int(_to_int_like(at(row, "rent"))),
                    "fee":     clean_int(_to_int_like(at(row, "fee"))),
                    "parking": clean_int(_to_int_like(at(row, "parking"))),
                    "water":   clean_int(_to_int_like(at(row, "water"))),
                    "reikin":  clean_int(_to_int_like(at(row, "reikin"))),
                    "koushin": clean_int(_to_int_like(at(row, "koushin"))),
                    "bikou":   at(row, "bikou"),
                }
            },
            "shikikin":    0,
            "linked_room": "",
        }
        records.append(rec)
    return records

async def handle_file_pdf(file_name: str, file_bytes: bytes, default_month_id: str):
    table = extract_income_table_with_pdfplumber(file_bytes)
    if not table:
        st.error(f"{file_name}: 収入明細の表を検出できませんでした（非AIルート）。")
        logger.error(f"{file_name}: pdfplumber抽出失敗（文字PDFでない/表検出不可/非空率低）")
        return []
    try:
        records = parse_income_table_to_records(table, default_month_id)
        if not records:
            st.error(f"{file_name}: 表は見つかりましたが、明細のパースに失敗しました（非AIルート）。")
            logger.error(f"{file_name}: テーブル→records 変換に失敗（列マッピング/数値化の不一致）")
            return []
        logger.info(f"{file_name}: 非AI解析成功 / {len(records)}件")
        return records
    except Exception as e:
        st.error(f"{file_name}: 明細のパースで例外が発生しました（非AIルート）。")
        logger.exception(e)
        return []

# =========================================================
# AIルート（画像→JSON→正規化）
# =========================================================
# OpenAI クライアントは必要になった時だけ初期化（Secrets未設定ならメッセージ）
_openai_client = None
def _get_openai_client():
    global _openai_client
    if _openai_client is None:
        try:
            from openai import AsyncOpenAI
            key = st.secrets.get("OPENAI_API_KEY", None)
            if not key:
                raise RuntimeError("OPENAI_API_KEY が設定されていません。")
            _openai_client = AsyncOpenAI(api_key=key)
        except Exception as e:
            raise RuntimeError(f"OpenAIクライアント初期化に失敗: {e}")
    return _openai_client

VISION_INSTRUCTIONS = (
    "あなたは不動産管理アシスタントです。収支報告書（送金明細書）から、"
    "各『室番号×賃借人（契約）』の入居情報を抽出し、厳格な JSON で返してください。\n"
    "要件:\n"
    "1) 出力は必ず次のトップレベル構造:\n"
    "{\n"
    "  \"records\": [\n"
    "    {\n"
    "      \"room\": \"0101\" または \"P01\" など,\n"
    "      \"tenant\": \"賃借人名\"（駐車場(Pxx)は空文字でも可）, \n"
    "      \"monthly\": {\n"
    "        \"YYYY-MM\": {\n"
    "          \"rent\": 家賃, \"fee\": 共益費, \"parking\": 駐車料, \"water\": 水道料,\n"
    "          \"reikin\": 礼金, \"koushin\": 更新料, \"bikou\": \"備考文字列\"\n"
    "        }, ...\n"
    "      },\n"
    "      \"shikikin\": 敷金合計（分かれば。なければ0）, \n"
    "      \"linked_room\": \"0001\" のように、Pxx行が特定住戸に紐付く場合に記す（備考の（0001）表記等から判断）\n"
    "    }, ...\n"
    "  ]\n"
    "}\n"
    "2) 各数値はカンマ無しの整数。空欄は 0。\n"
    "3) 月キーは YYYY-MM（例: 2024-11）。表に現れた全ての月を対象。\n"
    "4) 『P01/P02…』など駐車場の行は必ず room に Pxx を入れ、備考に「（0001）込駐車場」等があれば linked_room に『0001』のように数字4桁で格納。\n"
    "5) 同一室で入退去がある場合は賃借人ごとに別レコード（records の要素を分ける）。\n"
    "6) JSON 以外の文字（前置き・コードブロック）は出力しない。"
)


def convert_pdf_to_images(pdf_bytes, dpi=220):
    pdf = fitz.open(stream=pdf_bytes, filetype="pdf")
    images = []
    # 画像サイズの安定化（Matrix指定）
    scale = dpi / 72.0
    mat = fitz.Matrix(scale, scale)
    idx = 1 if len(pdf) >= 2 else 0 # 2ページ目だけから
    p = pdf[idx]
    pix = p.get_pixmap(matrix=mat, alpha=False)
    images.append(Image.frombytes("RGB", [pix.width, pix.height], pix.samples))
    return images


def convert_image_to_base64(image):
    buf = io.BytesIO()
    image.save(buf, format="JPEG", quality=90)
    return base64.b64encode(buf.getvalue()).decode("utf-8")

async def call_openai_vision_async(base64_images, default_month_id):
    client = _get_openai_client()
    image_parts = [{"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64}"}} for b64 in base64_images]
    messages = [
        {"role": "system", "content": VISION_INSTRUCTIONS},
        {"role": "user", "content": [
            *image_parts,
            {"type": "text", "text":
                f"このPDFには {default_month_id} 付近の月が含まれる可能性があります。"
                f"表内に現れた全ての『年／月』を抽出してください。\n\n"
                "※ 出力は純粋な JSON オブジェクトのみ。"}
        ]}
    ]

    resp = await client.chat.completions.create(
        model="gpt-4o",
        messages=messages,
        temperature=0.0,
        max_tokens=4096,
        response_format={"type": "json_object"},
    )

    return resp.choices[0].message.content

async def handle_file_ai(file_name: str, file_bytes: bytes, default_month_id: str, max_attempts: int = 3):
    images = convert_pdf_to_images(file_bytes)
    b64s = [convert_image_to_base64(img) for img in images]

    last_err = None
    for attempt in range(1, max_attempts + 1):
        try:
            raw = await call_openai_vision_async(b64s, default_month_id)
            s = raw.strip()
            s = s.removeprefix("```json").removeprefix("```").removesuffix("```").strip()
            obj = json.loads(s)
            if not isinstance(obj, dict) or "records" not in obj or not isinstance(obj["records"], list):
                raise ValueError("JSON ルートが {'records': [...]} になっていません。")
            norm_records = []
            for r in obj["records"]:
                room = normalize_room(r.get("room", ""))
                tenant = (r.get("tenant") or "").strip()
                shikikin = clean_int(r.get("shikikin"))
                linked_room = normalize_room(r.get("linked_room", "")) if r.get("linked_room") else ""
                monthly = {}
                for mk, mv in (r.get("monthly") or {}).items():
                    mk2 = month_key(mk)
                    monthly[mk2] = {
                        "rent":      clean_int((mv or {}).get("rent")),
                        "fee":       clean_int((mv or {}).get("fee")),
                        "parking":   clean_int((mv or {}).get("parking")),
                        "water":     clean_int((mv or {}).get("water")),
                        "reikin":    clean_int((mv or {}).get("reikin")),
                        "koushin":   clean_int((mv or {}).get("koushin")),
                        "bikou":     str((mv or {}).get("bikou") or "").strip(),
                    }
                norm_records.append({
                    "room": room, "tenant": tenant, "monthly": monthly,
                    "shikikin": shikikin, "linked_room": linked_room
                })
            logger.info(f"{file_name}: AI解析成功 / {len(norm_records)}件")
            return norm_records
        except Exception as e:
            last_err = e
            logger.warning(f"{file_name}: AI解析失敗（{attempt}/{max_attempts}）: {e}")
    st.warning(f"{file_name}: AI結果をJSONとして解釈できませんでした。")
    logger.error(f"{file_name}: 失敗の最終原因: {last_err}")
    return []

# =========================================================
# 入口：自動判定してルート切替
# =========================================================
async def handle_file(file):
    file_name = file.name
    logger.info(f"開始: {file_name}")
    default_month_id = extract_month_from_filename(file_name)
    file_bytes = file.read()

    if is_text_pdf(file_bytes):
        st.info(f"{file_name}: 文字情報を検出 → **非AI（高速・無料）** で解析します。")
        logger.info(f"{file_name}: route=non_ai")
        return await handle_file_pdf(file_name, file_bytes, default_month_id)
    else:
        st.info(f"{file_name}: 文字情報が見つからず → **AI** で解析します（料金/時間が発生）。")
        logger.info(f"{file_name}: route=ai")
        try:
            return await handle_file_ai(file_name, file_bytes, default_month_id)
        except RuntimeError as e:
            # OPENAI_API_KEY 未設定など
            st.error(f"{file_name}: AI解析に必要な設定がありません。{e}")
            logger.error(f"{file_name}: AIルート初期化失敗: {e}")
            return []

# ========== 全ファイル並列処理 & マージ ==========
def merge_records(all_recs, new_recs):
    for r in new_recs:
        key = (r["room"], r["tenant"])
        if key not in all_recs:
            all_recs[key] = {
                "room": r["room"],
                "tenant": r["tenant"],
                "monthly": {},
                "shikikin": clean_int(r.get("shikikin", 0)),
                "linked_room": r.get("linked_room", ""),
            }
        all_recs[key]["shikikin"] = max(all_recs[key]["shikikin"], clean_int(r.get("shikikin", 0)))
        for mk, mv in (r.get("monthly") or {}).items():
            dst = all_recs[key]["monthly"].setdefault(mk, {
                "rent":0,"fee":0,"parking":0,"water":0,"reikin":0,"koushin":0,"bikou":""
            })
            dst["rent"]    += clean_int(mv.get("rent"))
            dst["fee"]     += clean_int(mv.get("fee"))
            dst["parking"] += clean_int(mv.get("parking"))
            dst["water"]   += clean_int(mv.get("water"))
            dst["reikin"]  += clean_int(mv.get("reikin"))
            dst["koushin"] += clean_int(mv.get("koushin"))
            b = str(mv.get("bikou") or "").strip()
            if b:
                dst["bikou"] = append_note_unique(dst.get("bikou"), b)


def fold_parking_Pxx(all_recs):
    """
    Pxx レコードを、linked_room に駐車料として付替える。
    付替え先キーは (linked_room, tenant='') が見つからない場合は
    同室の誰かのレコード（賃借人がいるもの）にまとめる（最初に見つかったもの）。
    """
    to_delete = []
    # 検索用: room -> keys(list)
    by_room = {}
    for key, rec in all_recs.items():
        by_room.setdefault(rec["room"], []).append(key)

    for key, rec in list(all_recs.items()):
        room = rec["room"]
        if not room.upper().startswith("P"):
            continue
        # 付替え先
        target_room = rec.get("linked_room") or ""
        if not target_room:
            # 備考から (dddd) を拾う fallback
            for mk, mv in rec.get("monthly", {}).items():
                m = re.search(r"（?(\d{3,4})）?", mv.get("bikou",""))
                if m:
                    target_room = m.group(1).zfill(4)
                    break
        if not target_room:
            # 付替え不能なら残す（稀ケース）
            logger.info(f"Pxx行 {key} は付替え先不明のため残存")
            continue

        # 候補キー
        target_keys = by_room.get(target_room, [])
        if not target_keys:
            # まだ同室のレコードがない場合、空テナントのレコードを新設
            tkey = (target_room, "")
            all_recs[tkey] = {"room": target_room, "tenant":"", "monthly": {}, "shikikin":0, "linked_room":""}
            by_room.setdefault(target_room, []).append(tkey)
            target_keys = [tkey]

        # 付替えは最初の候補へ
        tkey = target_keys[0]
        target = all_recs[tkey]
        for mk, mv in rec.get("monthly", {}).items():
            dst = target["monthly"].setdefault(mk, {"rent":0,"fee":0,"parking":0,"water":0,"reikin":0,"koushin":0,"bikou":""})
            dst["parking"] += clean_int(mv.get("parking"))
            # 備考に「(P01→0001付替)」をメモ（任意）
            note = f"駐車場({room})→{target_room}"
            if note not in (dst["bikou"] or ""):
                dst["bikou"] = (dst["bikou"] + ", " if dst["bikou"] else "") + note

        to_delete.append(key)

    for key in to_delete:
        all_recs.pop(key, None)


def most_frequent_amount(values):
    """
    優先1: 非ゼロの最頻値
    優先2: 0を含めた最頻値
    優先3: 空なら 0
    最頻値が複数ある場合は、その中で最大値
    """
    vals = [clean_int(v) for v in values]
    def pick_mode(nums):
        if not nums:
            return None
        cnt = Counter(nums)
        max_freq = max(cnt.values())
        candidates = [v for v, f in cnt.items() if f == max_freq]
        return max(candidates)
    nonzero = [v for v in vals if v != 0]
    mode = pick_mode(nonzero)
    if mode is not None: return mode
    mode = pick_mode(vals)
    if mode is not None: return mode
    return 0

async def process_files(files):
    tasks = [handle_file(file) for file in files]
    results = await asyncio.gather(*tasks)

    # 1) マージ
    all_recs = {}
    for recs in results:
        merge_records(all_recs, recs)

    # 2) Pxx 付替え（契約者名で居室へ）
    fold_parking_Pxx(all_recs)

    # 3) 出力用に並べ替え & 基準額（最頻値）付与
    out = []
    for (room, tenant), rec in all_recs.items():
        def collect(k):
            return [clean_int(v.get(k,0)) for v in rec["monthly"].values()]
        rec["base_rent"]    = most_frequent_amount(collect("rent"))
        rec["base_fee"]     = most_frequent_amount(collect("fee"))
        rec["base_parking"] = most_frequent_amount(collect("parking"))
        rec["base_water"]   = most_frequent_amount(collect("water"))
        out.append(rec)

    def room_sort_key(r):
        rm = r["room"]
        if isinstance(rm, str) and rm.upper().startswith("P"):
            num = 9000 + int(re.sub(r"\D","",rm) or 0)
        else:
            num = int(re.sub(r"\D","",rm) or 0) if re.sub(r"\D","",str(rm)) else 9999
        first_month = sorted(r["monthly"].keys())[0] if r["monthly"] else "9999-99"
        return (num, r["tenant"] or "~", first_month)

    out_sorted = sorted(out, key=room_sort_key)
    months = sorted({m for r in out_sorted for m in r["monthly"].keys()})
    return out_sorted, months

# ========== Excel 生成 ==========
def combine_bikou_contract(rec):
    s = set()
    for mv in rec.get("monthly", {}).values():
        b = (mv.get("bikou") or "").strip()
        if b: s.add(b)
    return ", ".join(sorted(s))

def export_excel(records, months, property_name):
    wb = Workbook()
    ws = wb.active
    ws.title = xls_clean(property_name) or "入居管理表"

    header_row = 6
    data_start_row = 7
    last_fixed_col = 3
    number_fmt  = "#,##0"

    header_fill = PatternFill("solid", fgColor="BDD7EE")
    green_fill  = PatternFill("solid", fgColor="CCFFCC")
    gray_fill   = PatternFill("solid", fgColor="DDDDDD")
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_vert = Alignment(vertical="center", wrap_text=True)
    bold_font   = Font(bold=True)
    red_font    = Font(color="9C0000")
    thin_border = Border(*[Side(style='thin')] * 4)

    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    pink_fill   = PatternFill("solid", fgColor="F8CBAD")
    thick_side  = Side(style="thick")
    thick_border = Border(left=thick_side, right=thick_side, top=thick_side, bottom=thick_side)
    
    num_months = len(months)
    col_B = 2; col_C = 3; col_D = 4; col_E = 5; col_F = 6; col_G = 7
    col_month_end = 6 + num_months
    col_S = col_month_end + 1
    col_T = col_month_end + 2
    col_U = col_month_end + 3
    col_V = col_month_end + 4
    col_W = col_month_end + 5
    col_X = col_W + 1

    ws.merge_cells(start_row=2, start_column=col_B, end_row=2, end_column=col_W)
    if months:
        start_month = months[0].replace("-", "年") + "月"
        end_month   = months[-1].replace("-", "年") + "月"
        title_val = f"入居管理表 （{start_month}〜{end_month}）"
    else:
        title_val = "入居管理表"
    ws.cell(row=2, column=col_B, value=xls_clean(title_val)).font = Font(size=14, bold=True)
    ws.cell(row=2, column=col_B).alignment = center

    ws.merge_cells(start_row=4, start_column=col_B, end_row=4, end_column=col_C)
    ws.cell(row=4, column=col_B, value="物件名").alignment = center
    ws.merge_cells(start_row=4, start_column=col_D, end_row=4, end_column=col_F)
    ws.cell(row=4, column=col_D, value=xls_clean(property_name or "")).alignment = center
    for c in range(col_B, col_C+1):
        ws.cell(row=4, column=c).border = thick_border
    for c in range(col_D, col_F+1):
        ws.cell(row=4, column=c).border = thick_border

    ws.merge_cells(start_row=header_row, start_column=col_B, end_row=header_row, end_column=col_C)
    ws.cell(row=header_row, column=col_B, value="賃借人")
    ws.merge_cells(start_row=header_row, start_column=col_D, end_row=header_row, end_column=col_E)
    ws.cell(row=header_row, column=col_D, value="基準額")
    ws.cell(row=header_row, column=col_F, value="期首\n未収/前受")
    for i, m in enumerate(months):
        mm = int(m[5:])
        ws.cell(row=header_row, column=col_G+i, value=f"{mm}月")
    ws.cell(row=header_row, column=col_S, value="合計")
    ws.cell(row=header_row, column=col_T, value="期末\n未収/前受")
    ws.cell(row=header_row, column=col_U, value="礼金・更新料")
    ws.cell(row=header_row, column=col_V, value="敷金")
    ws.cell(row=header_row, column=col_W, value="備考")
    for c in range(col_B, col_W+1):
        cc = ws.cell(row=header_row, column=c)
        cc.fill = header_fill
        cc.font = bold_font
        cc.alignment = center

    # ---- データ（5行ブロック）----
    row = data_start_row
    for rec in records:
        room   = rec.get("room","")
        tenant = rec.get("tenant","")
        base_r = rec.get("base_rent",0)
        base_f = rec.get("base_fee",0)
        base_p = rec.get("base_parking",0)
        base_w = rec.get("base_water",0)
        shikikin = rec.get("shikikin",0)
        reikin_koushin_total = sum((mv.get("reikin",0)+mv.get("koushin",0)) for mv in rec.get("monthly",{}).values())

        ws.merge_cells(start_row=row,   start_column=col_B, end_row=row+4, end_column=col_B)
        ws.cell(row=row, column=col_B, value="室番号").alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        ws.cell(row=row, column=col_C, value=xls_clean(room)).alignment = center
        ws.cell(row=row, column=col_C).fill = green_fill
        ws.merge_cells(start_row=row+1, start_column=col_C, end_row=row+4, end_column=col_C)
        ws.cell(row=row+1, column=col_C, value=xls_clean(tenant)).alignment = center

        subjects = ["家賃","共益費　","駐車料","水道料","合計"]
        for i, s in enumerate(subjects):
            ws.cell(row=row+i, column=col_D, value=xls_clean(s))
        for i, v in enumerate([base_r, base_f, base_p, base_w]):
            cc = ws.cell(row=row+i, column=col_E, value=v); cc.number_format = number_fmt
        ws.cell(row=row+4, column=col_E, value=f"=SUM(E{row}:E{row+3})").number_format = number_fmt

        for i in range(5):
            ws.cell(row=row+i, column=col_F, value=0).number_format = number_fmt
        ws.cell(row=row+4, column=col_F, value=f"=SUM(F{row}:F{row+3})").number_format = number_fmt

        for i, m in enumerate(months):
            mv = (rec.get("monthly") or {}).get(m, {})
            vals = [mv.get("rent",0), mv.get("fee",0), mv.get("parking",0), mv.get("water",0)]
            for r_i, v in enumerate(vals):
                cc = ws.cell(row=row+r_i, column=col_G+i, value=v)
                cc.number_format = number_fmt
            ws.cell(row=row+4, column=col_G+i, value=f"=SUM({get_column_letter(col_G+i)}{row}:{get_column_letter(col_G+i)}{row+3})").number_format = number_fmt

        for r_i in range(5):
            ws.cell(row=row+r_i, column=col_S, value=f"=SUM({get_column_letter(col_G)}{row+r_i}:{get_column_letter(col_month_end)}{row+r_i})").number_format = number_fmt

        for r_i in range(4):
            ws.cell(row=row+r_i, column=col_T, value=0).number_format = number_fmt
        ws.cell(row=row+4, column=col_T, value=f"=SUM({get_column_letter(col_T)}{row}:{get_column_letter(col_T)}{row+3})").number_format = number_fmt

        ws.merge_cells(start_row=row, start_column=col_U, end_row=row+4, end_column=col_U)
        cu = ws.cell(row=row, column=col_U, value=reikin_koushin_total); cu.alignment = center_vert; cu.number_format = number_fmt
        ws.merge_cells(start_row=row, start_column=col_V, end_row=row+4, end_column=col_V)
        cv = ws.cell(row=row, column=col_V, value=shikikin); cv.alignment = center_vert; cv.number_format = number_fmt
        ws.merge_cells(start_row=row, start_column=col_W, end_row=row+4, end_column=col_W)
        bw = ws.cell(row=row, column=col_W, value=xls_clean(combine_bikou_contract(rec))); bw.alignment = center_vert; bw.font = red_font

        for c in range(col_B, col_W+1):
            for r in range(row, row+5):
                ws.cell(row=r, column=c).border = thin_border
        for c in range(col_B, col_W+1):
            ws.cell(row=row+4, column=c).fill = yellow_fill

        row += 5

    first_data_row = data_start_row
    last_data_row  = row - 1

    sum_start = row
    ws.merge_cells(start_row=sum_start, end_row=sum_start+3, start_column=col_B, end_column=col_C)
    ws.cell(row=sum_start, column=col_B, value="合計").alignment = center
    for i, name in enumerate(["家賃","共益費　","駐車料","水道料"]):
        ws.cell(row=sum_start+i, column=col_D, value=xls_clean(name))
    def sumif_range(col_letter):
        return f"{col_letter}${first_data_row}:{col_letter}${last_data_row}"
    for i in range(4):
        r = sum_start + i
        for cidx in range(col_E, col_T+1):
            col_letter = get_column_letter(cidx)
            ws.cell(row=r, column=cidx, value=f"=SUMIF($D${first_data_row}:$D${last_data_row},$D${r},{sumif_range(col_letter)})").number_format = number_fmt
    for cidx in [col_U, col_V]:
        col_letter = get_column_letter(cidx)
        ws.cell(row=sum_start, column=cidx, value=f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})").number_format = number_fmt
        for i in range(1,4):
            ws.cell(row=sum_start+i, column=cidx, value=None)
    for i in range(4):
        ws.cell(row=sum_start+i, column=col_W, value="")

    for c in range(col_B, col_W+1):
        for r in range(sum_start, sum_start+4):
            ws.cell(row=r, column=c).border = thin_border

    grand_row = sum_start + 4
    ws.merge_cells(start_row=grand_row, end_row=grand_row, start_column=col_B, end_column=col_C)
    ws.cell(row=grand_row, column=col_B, value="総合計").alignment = center
    for cidx in range(col_E, col_T+1):
        col_letter = get_column_letter(cidx)
        ws.cell(row=grand_row, column=cidx, value=f"=SUM({col_letter}{sum_start}:{col_letter}{sum_start+3})").number_format = number_fmt
    for c in range(col_B, col_W+1):
        ws.cell(row=grand_row, column=c).border = thin_border
        ws.cell(row=grand_row, column=c).fill = pink_fill

    ws.cell(row=grand_row-1, column=col_X, value=xls_clean("確認用")).alignment = center
    g_letter = get_column_letter(col_G); r_letter = get_column_letter(col_month_end)
    ws.cell(row=grand_row, column=col_X, value=f"=SUM({g_letter}{first_data_row}:{r_letter}{last_data_row})/2").number_format = number_fmt

    check_row = grand_row + 2
    ws.cell(row=check_row, column=col_E, value=xls_clean("算式確認"))
    for cidx in range(col_F, col_T+1):
        col_letter = get_column_letter(cidx)
        ws.cell(row=check_row, column=cidx, value=f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})/2").number_format = number_fmt

    ws.column_dimensions[get_column_letter(col_W)].width = max(
        [len(xls_clean(combine_bikou_contract(rec)) or "") for rec in records] + [10]
    ) * 1.6

    try:
        ws.freeze_panes = ws.cell(row=data_start_row, column=last_fixed_col+1)  # "D7"
    except Exception:
        pass

    # 開いたときにフル再計算（人が開く時）
    wb.calculation.fullCalcOnLoad = True

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

# ========== Streamlit UI ==========
st.set_page_config(page_title="入居管理表アプリ（自動切替）", layout="wide")
st.title("📊 収支報告書PDFから入居管理表を作成（自動切替）")

PASSWORD = st.secrets["APP_PASSWORD"]
pw = st.text_input("パスワードを入力してください", type="password")
if pw != PASSWORD:
    st.warning("正しいパスワードを入力してください。")
    st.stop()

property_name = st.text_input("物件名（例：XOヒルズ）", value="")
uploaded_files = st.file_uploader("収支報告書PDFを最大12ファイルまでアップロードしてください", type="pdf", accept_multiple_files=True)

if uploaded_files and st.button("入居管理表を作成"):
    if len(uploaded_files) > 12:
        st.warning("アップロードできるのは最大12ファイルまでです。")
    else:
        st.info("収支報告書を読み取り中…（文字PDFなら非AI／スキャンPDFならAIで解析）")
        records, months = asyncio.run(process_files(uploaded_files))
        if not records:
            st.error("データが抽出できませんでした。PDFの種類・品質をご確認ください。")
            st.stop()

        st.info("入居管理表を作成中...")
        excel_data = export_excel(records, months, property_name)
        if months:
            start_month = months[0].replace("-", "年") + "月"
            end_month   = months[-1].replace("-", "年") + "月"
            fn = f"{xls_clean(property_name) or '入居管理表'}（{start_month}〜{end_month}）_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx"
        else:
            fn = f"{xls_clean(property_name) or '入居管理表'}_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx"

        st.download_button("入居管理表をダウンロード", data=excel_data,
                           file_name=fn,
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        st.success("完了しました。")

